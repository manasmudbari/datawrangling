# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import openpyxl
import os

os.chdir('/Users/manasmudbari/Desktop')

wb = openpyxl.load_workbook('projects.xlsx')

sheet = wb.get_sheet_by_name('Single Author')

respectCnt = 0
commentsCnt = 0
impCnt = 0

for i in range (2,916):
    if sheet.cell(row=i, column=7).value == sheet.cell(row=i-1, column=7).value:
        respectCnt = respectCnt + sheet.cell(row=i, column=5).value
    else:
        newcnt = respectCnt        
        respectCnt = sheet.cell(row=i, column=5).value
        print newcnt

for j in range (2,916):
    if sheet.cell(row=j, column=7).value == sheet.cell(row=j-1, column=7).value:
        commentsCnt = commentsCnt + sheet.cell(row=j, column=6).value
    else:
        newcmt = commentsCnt
        commentsCnt = sheet.cell(row=j, column=6).value
        print newcmt

for k in range (2,916):
    if sheet.cell(row=k, column=7).value == sheet.cell(row=k-1, column=7).value:
        impCnt = impCnt + sheet.cell(row=k, column=4).value
    else:
        newimp = impCnt
        impCnt = sheet.cell(row=k, column=4).value
        print newimp